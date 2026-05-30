"""
Загрузка демо-данных из каталога test_files в корне репозитория.

Используются файлы:
  - Tovar.xlsx           — товары и фото (1.jpg … 20.jpg)
  - user_import.xlsx     — пользователи
  - Заказ_import.xlsx    — заказы
  - Пункты выдачи_import.xlsx — адреса пунктов выдачи
"""

from __future__ import annotations

import shutil
import sqlite3
from datetime import datetime
import calendar
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

ROOT_DIR = Path(__file__).resolve().parent.parent.parent
TEST_FILES_DIR = ROOT_DIR / "test_files"
DB_PATH = Path(__file__).resolve().parent.parent / "storage" / "database" / "book_store_exam.db"
IMAGES_DIR = Path(__file__).resolve().parent.parent / "storage" / "images"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

ROLE_MAP = {
    "Администратор": "admin",
    "Менеджер": "manager",
    "Авторизованный клиент": "client",
}


def find_xlsx(name_part: str, *, exclude: str | None = None) -> Path:
    matches = sorted(
        path
        for path in TEST_FILES_DIR.glob(f"*{name_part}*.xlsx")
        if exclude is None or exclude.lower() not in path.name.lower()
    )
    if not matches:
        raise FileNotFoundError(f"XLSX not found in {TEST_FILES_DIR}: *{name_part}*.xlsx")
    return matches[0]


def load_pickup_points() -> list[str]:
    path = find_xlsx("выдачи")
    workbook = openpyxl.load_workbook(path, read_only=True)
    try:
        sheet = workbook.active
        return [row[0] for row in sheet.iter_rows(values_only=True) if row[0]]
    finally:
        workbook.close()


def load_users() -> list[tuple]:
    path = find_xlsx("user_import")
    workbook = openpyxl.load_workbook(path, read_only=True)
    try:
        sheet = workbook.active
        rows = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            role_raw, full_name, login, password = row[:4]
            if not login or not password:
                continue
            role = ROLE_MAP.get(str(role_raw).strip(), "client")
            rows.append((index, str(full_name).strip(), str(login).strip(), str(password).strip(), role, 1))
        return rows
    finally:
        workbook.close()


def copy_product_image(photo_name: str | None) -> str | None:
    if not photo_name:
        return None
    source = TEST_FILES_DIR / str(photo_name).strip()
    if not source.exists():
        return None
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    target = IMAGES_DIR / source.name
    shutil.copy2(source, target)
    return target.name


def load_products() -> list[tuple]:
    path = TEST_FILES_DIR / "Tovar.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Products file not found: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True)
    try:
        sheet = workbook.active
        rows = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            if not row or not row[0]:
                continue
            (
                article,
                name,
                unit,
                price,
                supplier,
                manufacturer,
                category,
                discount,
                stock,
                description,
                photo,
            ) = row[:11]
            image_path = copy_product_image(photo)
            rows.append(
                (
                    index,
                    str(article).strip(),
                    str(name).strip(),
                    str(category).strip() if category else "",
                    str(description).strip() if description else "",
                    str(manufacturer).strip() if manufacturer else "",
                    str(supplier).strip() if supplier else "",
                    str(unit).strip() if unit else "шт.",
                    float(price or 0),
                    int(stock or 0),
                    float(discount or 0),
                    image_path,
                    NOW,
                    NOW,
                )
            )
        return rows
    finally:
        workbook.close()


def format_datetime(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue

    parts = text.split(".")
    if len(parts) == 3 and all(part.isdigit() for part in parts):
        day, month, year = (int(parts[0]), int(parts[1]), int(parts[2]))
        last_day = calendar.monthrange(year, month)[1]
        safe_day = min(day, last_day)
        parsed = datetime(year, month, safe_day)
        return parsed.strftime("%Y-%m-%d %H:%M:%S")

    return NOW


def parse_order_items(raw: str) -> list[tuple[str, int]]:
    parts = [part.strip() for part in str(raw).split(",") if part.strip()]
    items: list[tuple[str, int]] = []
    index = 0
    while index + 1 < len(parts):
        article = parts[index]
        quantity = int(float(parts[index + 1]))
        items.append((article, quantity))
        index += 2
    return items


def load_orders(article_to_product: dict[str, tuple], manager_id: int) -> list[tuple]:
    path = find_xlsx("Заказ", exclude="выдачи")
    pickup_points = load_pickup_points()
    workbook = openpyxl.load_workbook(path, read_only=True)
    order_rows: list[tuple] = []
    order_id = 1

    try:
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue

            (
                _order_number,
                articles_raw,
                order_date,
                delivery_date,
                pickup_index,
                customer_name,
                pickup_code,
                status,
            ) = row[:8]

            pickup_address = ""
            if isinstance(pickup_index, int) and 1 <= pickup_index <= len(pickup_points):
                pickup_address = str(pickup_points[pickup_index - 1])

            order_date_text = format_datetime(order_date) or NOW
            delivery_date_text = format_datetime(delivery_date)

            for article, quantity in parse_order_items(articles_raw):
                product = article_to_product.get(article)
                if product is None:
                    continue

                product_id, unit_price, discount_percent = product
                order_rows.append(
                    (
                        order_id,
                        str(customer_name).strip() if customer_name else "Без клиента",
                        manager_id,
                        str(status).strip() if status else "Новый",
                        pickup_address,
                        order_date_text,
                        delivery_date_text,
                        str(pickup_code).strip() if pickup_code else "",
                        "",
                        product_id,
                        quantity,
                        unit_price,
                        discount_percent,
                    )
                )
                order_id += 1
        return order_rows
    finally:
        workbook.close()


def main() -> None:
    if not TEST_FILES_DIR.exists():
        raise SystemExit(f"Test files directory not found: {TEST_FILES_DIR}")
    if not DB_PATH.exists():
        raise SystemExit(f"Database not found: {DB_PATH}. Start backend once to create schema.")

    users = load_users()
    products = load_products()
    if not users:
        raise SystemExit("No users found in user_import.xlsx")
    if not products:
        raise SystemExit("No products found in Tovar.xlsx")

    article_to_product = {
        row[1]: (row[0], row[8], row[10])
        for row in products
    }
    manager_id = next((user[0] for user in users if user[4] == "manager"), users[0][0])
    orders = load_orders(article_to_product, manager_id)

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("DELETE FROM orders")
        conn.execute("DELETE FROM products")
        conn.execute("DELETE FROM users")

        conn.executemany(
            "INSERT INTO users (id, full_name, login, password, role, active) VALUES (?, ?, ?, ?, ?, ?)",
            users,
        )
        conn.executemany(
            """
            INSERT INTO products (
                id, article, name, category, description, manufacturer, supplier, unit,
                price, stock_quantity, discount_percent, image_path, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            products,
        )
        if orders:
            conn.executemany(
                """
                INSERT INTO orders (
                    id, customer_name, manager_id, status, pickup_address, order_date, delivery_date,
                    pickup_code, comment, product_id, quantity, unit_price, discount_percent
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                orders,
            )

        conn.commit()
        print(f"Seed completed from {TEST_FILES_DIR}:")
        for table in ("users", "products", "orders"):
            count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            print(f"  {table}: {count}")
        images_count = len(list(IMAGES_DIR.glob("*"))) if IMAGES_DIR.exists() else 0
        print(f"  images copied: {images_count}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
